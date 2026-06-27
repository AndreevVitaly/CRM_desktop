from __future__ import annotations

from pathlib import Path

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from models.db_models import (
    DOCUMENT_TYPE_MEETING,
    Document,
    Encounter,
    Patient,
    PatientInteraction,
    User,
)
from ui.styles import FONTS, RADIUS, get_colors, get_main_stylesheet
from utils.office_export import (
    build_encounter_docx_filename,
    export_encounter_to_docx,
)


class EncountersPage(QWidget):
    def __init__(self, user: User):
        super().__init__()
        self.user = user
        self._visible_rows: list[dict] = []
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        filters = QFrame()
        filters.setObjectName("card")
        filter_layout = QHBoxLayout(filters)
        filter_layout.setContentsMargins(12, 12, 12, 12)
        filter_layout.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setObjectName("searchInput")
        self.search_input.setPlaceholderText(
            "Поиск по пациенту, номеру, врачу, документу или признаку"
        )
        self.search_input.setMinimumWidth(300)
        self.search_input.setFixedHeight(34)
        self.search_input.textChanged.connect(self._load_encounters)
        filter_layout.addWidget(self.search_input)

        self.status_filter = QComboBox()
        self.status_filter.setObjectName("filterCombo")
        self.status_filter.setFrame(False)
        self.status_filter.addItem("Все статусы", "")
        self.status_filter.addItem("Запланирован", Encounter.STATUS_PLANNED)
        self.status_filter.addItem("В процессе", Encounter.STATUS_INPROGRESS)
        self.status_filter.addItem("Завершен", Encounter.STATUS_FINISHED)
        self.status_filter.setFixedHeight(34)
        self.status_filter.setFixedWidth(160)
        self.status_filter.setStyleSheet(self._filter_combo_style())
        self.status_filter.currentIndexChanged.connect(self._load_encounters)
        filter_layout.addWidget(self.status_filter)

        filter_layout.addStretch()

        refresh_btn = self._action_button("Обновить")
        refresh_btn.clicked.connect(self._load_encounters)
        filter_layout.addWidget(refresh_btn)

        word_btn = self._action_button("WORD")
        word_btn.clicked.connect(self._export_selected_word)
        filter_layout.addWidget(word_btn)

        excel_btn = self._action_button("EXCEL")
        excel_btn.clicked.connect(self._export_visible_excel)
        filter_layout.addWidget(excel_btn)

        layout.addWidget(filters)

        self.count_label = QLabel()
        self.count_label.setObjectName("muted")
        layout.addWidget(self.count_label)

        self.table = QTableWidget()
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels(
            [
                "Дата",
                "Пациент",
                "Личный номер",
                "Врач",
                "Результат",
                "Причина",
                "Признак",
                "Статус",
                "Важность",
                "Информация",
                "Документ",
            ]
        )
        header = self.table.horizontalHeader()
        for column in (0, 1, 2, 3, 4, 6, 7, 8, 10):
            header.setSectionResizeMode(column, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.doubleClicked.connect(self._edit_selected_encounter)
        layout.addWidget(self.table, 1)

        self.setStyleSheet(get_main_stylesheet())
        self._load_encounters()

    def _action_button(self, text: str) -> QPushButton:
        button = QPushButton(text)
        button.setObjectName("filterButton")
        button.setFixedHeight(34)
        button.setMinimumWidth(92)
        button.setCursor(Qt.CursorShape.PointingHandCursor)
        button.setStyleSheet(self._filter_button_style())
        return button

    def _filter_combo_style(self) -> str:
        colors = get_colors()
        return f"""
            QComboBox#filterCombo {{
                background-color: {colors['surface']};
                border: 1px solid {colors['line']};
                border-radius: {RADIUS['sm']}px;
                padding: 4px 12px;
                color: {colors['text']};
                font-size: {FONTS['size_small']}pt;
                min-width: 0px;
            }}
            QComboBox#filterCombo:hover {{
                border: 1px solid {colors['accent']};
            }}
            QComboBox#filterCombo:focus {{
                border: 1px solid {colors['accent']};
                background-color: {colors['accent_light']};
            }}
            QComboBox#filterCombo::drop-down {{
                border: none;
                width: 0px;
            }}
            QComboBox#filterCombo::down-arrow {{
                image: none;
                width: 0px;
                height: 0px;
                border: none;
            }}
            QComboBox#filterCombo QAbstractItemView {{
                background-color: {colors['surface']};
                border: 1px solid {colors['line']};
                selection-background-color: {colors['accent_light']};
                selection-color: {colors['text']};
                outline: none;
            }}
        """

    def _filter_button_style(self) -> str:
        colors = get_colors()
        return f"""
            QPushButton#filterButton {{
                background-color: {colors['surface']};
                border: 1px solid {colors['line']};
                border-radius: {RADIUS['sm']}px;
                padding: 4px 12px;
                min-height: 0px;
                font-weight: 500;
                font-size: {FONTS['size_xs']}pt;
                color: {colors['text']};
            }}
            QPushButton#filterButton:hover {{
                background-color: {colors['accent_light']};
                border: 1px solid {colors['accent']};
                color: {colors['accent']};
            }}
            QPushButton#filterButton:pressed {{
                background-color: {colors['accent']};
                border: 1px solid {colors['accent']};
                color: #FFFFFF;
            }}
        """

    def _load_encounters(self):
        search = self.search_input.text().strip().casefold()
        status_filter = self.status_filter.currentData() or ""

        visible_rows = []
        for document in Document.get_all(self.user):
            if document.doc_type != DOCUMENT_TYPE_MEETING:
                continue

            patient = document.patient
            if not patient:
                continue

            encounter = document.encounter
            status = (
                encounter.status
                if encounter and encounter.status
                else Encounter.STATUS_FINISHED
            )
            if status_filter and status != status_filter:
                continue

            importance = (
                encounter.information_importance
                if encounter and encounter.information_importance
                else ""
            )

            doctor = (
                (encounter.doctor if encounter and encounter.doctor_id else None)
                or patient.doctor
                or document.author
            )
            doctor_name = doctor.full_name if doctor else "—"
            document_number = str(document.doc_number or f"#{document.id}")
            group = encounter.group if encounter and encounter.group_id else None
            group_name = group.name if group else ""
            haystack = " ".join(
                [
                    patient.callsign or "",
                    patient.personal_number or "",
                    doctor_name,
                    document_number,
                    document.summary or "",
                    group_name,
                ]
            ).casefold()
            if search and search not in haystack:
                continue

            visible_rows.append(
                {
                    "document": document,
                    "patient": patient,
                    "encounter": encounter,
                    "date": (
                        document.doc_date.strftime("%d.%m.%Y")
                        if document.doc_date
                        else "—"
                    ),
                    "doctor": doctor_name,
                    "result": (
                        encounter.meeting_result_display
                        if encounter and encounter.meeting_result
                        else "—"
                    ),
                    "reason": (
                        encounter.reason
                        if encounter and encounter.reason
                        else document.summary or "—"
                    ),
                    "group": group_name or "—",
                    "status": status,
                    "status_display": (
                        encounter.status_display if encounter else "Завершен"
                    ),
                    "importance": importance or "—",
                    "patient_info": (
                        encounter.patient_info if encounter else ""
                    )
                    or "—",
                    "document_number": document_number,
                }
            )

        self._visible_rows = visible_rows
        self._render_rows()

    def _render_rows(self):
        colors = get_colors()
        status_colors = {
            Encounter.STATUS_PLANNED: colors["accent"],
            Encounter.STATUS_INPROGRESS: colors["warning"],
            Encounter.STATUS_FINISHED: colors["success"],
        }
        self.table.setRowCount(0)
        for data in self._visible_rows:
            row = self.table.rowCount()
            self.table.insertRow(row)
            values = [
                data["date"],
                data["patient"].callsign or "—",
                data["patient"].personal_number or "—",
                data["doctor"],
                data["result"],
                data["reason"],
                data["group"],
                data["status_display"],
                data["importance"],
                data["patient_info"],
                data["document_number"],
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if column == 0:
                    item.setData(
                        Qt.ItemDataRole.UserRole, data["document"].id
                    )
                if column == 7:
                    item.setForeground(
                        QColor(status_colors.get(data["status"], colors["text"]))
                    )
                self.table.setItem(row, column, item)

        self.count_label.setText(f"Встреч: {len(self._visible_rows)}")

    def _selected_data(self) -> dict | None:
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            return None
        row = selected[0].row()
        if row < 0 or row >= len(self._visible_rows):
            return None
        return self._visible_rows[row]

    def _edit_selected_encounter(self, _index=None):
        data = self._selected_data()
        if not data:
            QMessageBox.warning(self, "Встречи", "Выберите встречу")
            return

        document = data["document"]
        patient = data["patient"]
        encounter = data["encounter"]
        if not encounter:
            encounter = Encounter(
                patient_id=patient.id,
                doctor_id=(
                    self.user.id if self.user.role == User.ROLE_DOCTOR else 0
                ),
                started_at=document.doc_date,
                reason=document.summary or "",
                status=Encounter.STATUS_FINISHED,
                document_id=document.id,
            )
            encounter.save()
            document.encounter_id = encounter.id
            document.save()

        from ui.encounter_edit_form import EncounterEditDialog

        dialog = EncounterEditDialog(self.user, patient, encounter)
        if dialog.exec():
            PatientInteraction(
                patient_id=patient.id,
                user_id=self.user.id,
                action="visit_edit",
                description=f"Отредактирована встреча (док. №{document.id})",
            ).save()
            self._load_encounters()

    def _export_selected_word(self):
        data = self._selected_data()
        if not data:
            QMessageBox.warning(self, "Экспорт Word", "Выберите встречу")
            return

        encounter = data["encounter"] or Encounter(
            patient_id=data["patient"].id,
            doctor_id=data["document"].author_id or 0,
            started_at=data["document"].doc_date,
            reason=data["document"].summary or "",
            status=Encounter.STATUS_FINISHED,
            document_id=data["document"].id,
        )
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить встречу в Word",
            build_encounter_docx_filename(data["patient"], encounter),
            "Word (*.docx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".docx"):
            file_path += ".docx"
        try:
            export_encounter_to_docx(data["patient"], encounter, file_path)
        except Exception as exc:
            QMessageBox.critical(
                self, "Экспорт Word", f"Не удалось создать файл:\n{exc}"
            )
            return
        QMessageBox.information(self, "Экспорт Word", "Встреча сохранена в Word")

    def _export_visible_excel(self):
        if not self._visible_rows:
            QMessageBox.information(self, "Экспорт Excel", "Нет встреч для экспорта")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить встречи в Excel",
            "pulsar_all_meetings.xlsx",
            "Excel (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Встречи"
            headers = [
                "Дата",
                "Пациент",
                "Личный номер",
                "Врач",
                "Результат",
                "Причина",
                "Признак",
                "Статус",
                "Важность",
                "Информация",
                "Документ",
            ]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            for data in self._visible_rows:
                sheet.append(
                    [
                        data["date"],
                        data["patient"].callsign or "—",
                        data["patient"].personal_number or "—",
                        data["doctor"],
                        data["result"],
                        data["reason"],
                        data["group"],
                        data["status_display"],
                        data["importance"],
                        data["patient_info"],
                        data["document_number"],
                    ]
                )
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            widths = [14, 18, 16, 24, 22, 34, 24, 16, 18, 42, 18]
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[
                    chr(64 + index)
                ].width = width
            sheet.freeze_panes = "A2"
            workbook.save(str(Path(file_path)))
        except Exception as exc:
            QMessageBox.critical(
                self, "Экспорт Excel", f"Не удалось создать файл:\n{exc}"
            )
            return
        QMessageBox.information(self, "Экспорт Excel", "Встречи сохранены в Excel")

    def update_theme(self):
        self.setStyleSheet(get_main_stylesheet())
        self.status_filter.setStyleSheet(self._filter_combo_style())
        for button in self.findChildren(QPushButton, "filterButton"):
            button.setStyleSheet(self._filter_button_style())
        self._render_rows()
